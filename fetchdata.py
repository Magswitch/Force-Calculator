import openpyxl  # Need this for working with excel docs

def getUnitData(unit, type):

	#Open our workbook and filter to the sheet coresponding to the name of the provided unit
	wb = openpyxl.load_workbook('New Master Test Collection Document.xlsx', data_only=True) #data_only to avoid excel refrences
	sheet = wb.get_sheet_by_name(unit)

	data = []

	#Depending on what data we want to return, either grab the forces or thicknesses
	if type == 'thicknesses':
		for i in range(4, sheet.max_row, 1):
			data.append(sheet.cell(row=i, column=2).value) #Column 2 is standard for thicknesses

	elif type == 'forces':
		for i in range(4, sheet.max_row, 1):
			data.append(sheet.cell(row=i, column=9).value) # Column 9 is standard for forces

	else:
		print("Unknown type of unit data %s" %type)

	return data