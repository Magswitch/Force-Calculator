import openpyxl

def getUnitData(unit):

	wb = openpyxl.load_workbook('New Master Test Collection Document.xlsx')
	sheet = wb.get_sheet_by_name(unit)

	data = []

	for i in range(3, 28, 1):
		data.append(sheet.cell(row=i, column=2).value, sheet.cell(row=i,column=3).value)

	return data